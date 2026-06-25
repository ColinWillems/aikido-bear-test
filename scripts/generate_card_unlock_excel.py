#!/usr/bin/env python3
"""Generate an Excel workbook listing every card with its plaintext ID,
obfuscated code and ready-to-share deep link.

The card list is the single source of truth shared with
``generate_unlock_qr_html.py`` (top of that file). One worksheet is written
per collection; within a sheet the cards are grouped by set with a colored
header row matching the in-app accent colour.

Columns per card:
    Set | Index | Sport | Card ID | Encoded ID | Deep Link

Output: bear_adventure_app/delivery/_firebase_upload/unlock_cards.xlsx

Prerequisites: the shared ``scripts/.venv`` virtualenv (see
``scripts/README.md``). ``openpyxl`` is listed in ``scripts/requirements.txt``.
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parent))
from card_catalogue import BASE_URL, CARDS, SET_COLORS  # noqa: E402
from card_id_obfuscator import encode_card_id  # noqa: E402


HEADERS: tuple[str, ...] = (
    "Set",
    "Index",
    "Sport",
    "Card ID",
    "Encoded ID",
    "Deep Link",
)

# Column widths (1-indexed to match the HEADERS tuple).
COL_WIDTHS: dict[int, int] = {
    1: 22,  # Set
    2: 8,   # Index
    3: 22,  # Sport
    4: 16,  # Card ID
    5: 14,  # Encoded ID
    6: 60,  # Deep Link
}

OUTPUT_PATH = Path(
    "bear_adventure_app/delivery/_firebase_upload/unlock_cards.xlsx"
)


def _safe_sheet_title(title: str) -> str:
    """Excel sheet titles are <= 31 chars and may not contain ``[]:*?/\\``."""
    cleaned = title
    for bad in "[]:*?/\\":
        cleaned = cleaned.replace(bad, "-")
    return cleaned[:31]


def _hex_to_argb(hex_color: str) -> str:
    """Convert ``#RRGGBB`` (or ``RRGGBB``) to openpyxl's ``AARRGGBB``."""
    h = hex_color.lstrip("#")
    if len(h) != 6:
        raise ValueError(f"expected 6-digit hex color, got {hex_color!r}")
    return f"FF{h.upper()}"


def _is_dark(hex_color: str) -> bool:
    """Rough luminance check so set-header text stays readable."""
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    # Rec. 709 luma
    return (0.2126 * r + 0.7152 * g + 0.0722 * b) < 140


def _write_header_row(ws: Worksheet, row: int) -> None:
    header_fill = PatternFill("solid", fgColor="FF2B1810")
    header_font = Font(bold=True, color="FFF5EBD8", size=11)
    for col_idx, label in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=False,
        )


def _write_set_header(ws: Worksheet, row: int, set_label: str, color: str) -> None:
    fill = PatternFill("solid", fgColor=_hex_to_argb(color))
    text_color = "FFFFFFFF" if _is_dark(color) else "FF2B1810"
    font = Font(bold=True, color=text_color, size=11)
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=len(HEADERS),
    )
    cell = ws.cell(row=row, column=1, value=set_label.upper())
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _write_card_row(
    ws: Worksheet,
    row: int,
    set_label: str,
    idx: int,
    sport: str,
    card_id: str,
    code: str,
    deep_link: str,
) -> None:
    values: tuple = (set_label, idx, sport, card_id, code, deep_link)
    mono = Font(name="Menlo", size=10)
    bold_mono = Font(name="Menlo", size=10, bold=True)
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col_idx == 4:           # Card ID
            cell.font = mono
        elif col_idx == 5:         # Encoded ID
            cell.font = bold_mono
        elif col_idx == 6:         # Deep Link
            cell.font = mono
            cell.hyperlink = deep_link
            cell.style = "Hyperlink"


def _apply_column_widths(ws: Worksheet) -> None:
    for col_idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _build_workbook() -> tuple[Workbook, int]:
    wb = Workbook()
    # Remove the default blank sheet; we add our own per collection.
    default_sheet = wb.active
    wb.remove(default_sheet)

    total_cards = 0

    for collection_title, sets in CARDS.items():
        ws = wb.create_sheet(title=_safe_sheet_title(collection_title))
        ws.sheet_view.showGridLines = False
        _apply_column_widths(ws)

        # Row 1: collection title banner spanning all columns.
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=len(HEADERS),
        )
        title_cell = ws.cell(row=1, column=1, value=collection_title)
        title_cell.font = Font(bold=True, size=16, color="FF2B1810")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # Row 2: column headers.
        _write_header_row(ws, row=2)
        ws.freeze_panes = "A3"

        row = 3
        for set_label, items in sets:
            color = SET_COLORS.get(set_label, "#888888")
            _write_set_header(ws, row=row, set_label=set_label, color=color)
            row += 1
            for card_id, sport, idx in items:
                code = encode_card_id(card_id)
                deep_link = f"{BASE_URL}/?path={code}"
                _write_card_row(
                    ws,
                    row=row,
                    set_label=set_label,
                    idx=idx,
                    sport=sport,
                    card_id=card_id,
                    code=code,
                    deep_link=deep_link,
                )
                total_cards += 1
                row += 1
            # Blank spacer row between sets.
            row += 1

    return wb, total_cards


def _iter_all_cards() -> Iterable[tuple[str, str, str, int, str]]:
    """Yield (collection_title, set_label, card_id, idx, sport) for sanity
    checks. Currently unused at runtime but handy when debugging."""
    for collection_title, sets in CARDS.items():
        for set_label, items in sets:
            for card_id, sport, idx in items:
                yield collection_title, set_label, card_id, idx, sport


def main() -> int:
    wb, total = _build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH} ({total} cards across "
          f"{len(CARDS)} collection(s))")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
